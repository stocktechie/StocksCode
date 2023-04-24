import concurrent.futures
import requests
from bs4 import BeautifulSoup
import openpyxl

# Fetch stock prices
file_path = '<PATH_TO_EXCEL>.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb['Sheet1']
max_row = sheet.max_row

base_url = 'https://www.google.com/finance/quote/{}?hl=en'
prices = {}

def fetch_price(symbol_exchange):
    symbol_exchange_split = symbol_exchange.split(":")
    symbol = symbol_exchange_split[1]
    exchange = symbol_exchange_split[0]
    url = base_url.format(symbol_exchange)
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    price_element = soup.find('div', {'class': 'YMlKec fxKbKc'})
    if price_element is not None:
        price = price_element.text
        print(f'{symbol_exchange}: {price}')
    else:
        print(f'Price not found for {symbol_exchange}')
        price = "Price not found"
    prices[symbol_exchange] = price

with concurrent.futures.ThreadPoolExecutor() as executor:
    futures = []
    for i in range(2, max_row + 1):
        symbol_exchange = sheet.cell(row=i, column=1).value
        if symbol_exchange is not None:
            futures.append(executor.submit(fetch_price, symbol_exchange))

    # Wait for all futures to complete
    concurrent.futures.wait(futures)

# Write stock prices to Excel file
for i in range(2, max_row + 1):
    symbol_exchange = sheet.cell(row=i, column=1).value
    cell = sheet.cell(row=i, column=2)
    cell.value = prices.get(symbol_exchange, "Price not found")

wb.save(file_path)
